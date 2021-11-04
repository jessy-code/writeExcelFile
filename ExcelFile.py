from openpyxl import Workbook, load_workbook
from os.path import isfile


class ExcelFile:
    def __init__(self, file_path):
        self.file_path = file_path
        self.excel_file = ''

        self.__call__()

    def __call__(self):
        self.excel_file = self.load_excel_file() if isfile(self.file_path) else self.create_empty_excel_file()

    def create_empty_excel_file(self):
        self.excel_file = Workbook()
        self.excel_file.save(self.file_path)
        return self.excel_file

    def load_excel_file(self):
        self.excel_file = load_workbook(self.file_path)
        return self.excel_file

    def write_in_cell_of_sheet(self, sheet, row, column, entry):
        sheet_target = self.excel_file[sheet]
        sheet_target.cell(row=row, column=column).value = entry
        self.excel_file.save(self.file_path)

    def write_list_in_line_of_sheet(self, sheet, line_number, list_entry):
        index = 1
        for entry in list_entry:
            self.write_in_cell_of_sheet(sheet, line_number, index, entry)
            index += 1
        self.excel_file.save(self.file_path)
