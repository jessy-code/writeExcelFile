from openpyxl import Workbook, load_workbook
from os.path import isfile


class ExcelFile:
    """
    This class allow to write in excel files using the openpyxl library.

    Attributes
    __________
    file_path : string
                The path where the excel file is stored

    excel_file : excel file modelling with a Workbook object from openpyxl library

    Parameters
    __________
    file_path : string
                The path where you want to write your excel file

    Methods
    _______
    create_empty_excel_file()
        Create an empty excel file in file_path and store it in excel_file attribute

    load_excel_file()
        Load an existing excel file and store it in excel_file attribute

    write_in_cell_of_sheet(sheet, row, column, entry)
        write the <entry> value in the row, column cell of the given sheet of the excel_file object

    write_list_in_line_of_sheet(self, sheet, line_number, list_entry, start_column=1)
        write the <list_entry> in the row, column cell of the given sheet of the excel_file object
    """
    def __init__(self, file_path):
        self.file_path = file_path
        self.excel_file = ''

        self.__call__()

    def __call__(self):
        """
        This function create or load the excel file to the instance creation.

        Raises
        ______
        FileNotFoundError
            If the file_path is impossible to use
        """
        try:
            self.excel_file = self.load_excel_file() if isfile(self.file_path) else self.create_empty_excel_file()
        except FileNotFoundError as e:
            print("Impossible to create or find " + self.file_path + ". Please check the file path.")
            raise

    def create_empty_excel_file(self):
        """
        Create an empty excel file in file_path and store it in excel_file attribute

        RETURNS
        _______
        self.excel_file
            the excel file object modelling with a Workbook object of openpyxl library
        """
        self.excel_file = Workbook()
        self.excel_file.save(self.file_path)
        return self.excel_file

    def load_excel_file(self):
        """
        Load an existing excel file and store it in excel_file attribute

        RETURNS
        _______
        self.excel_file
            the excel file object modelling with a Workbook object of openpyxl library
        """
        self.excel_file = load_workbook(self.file_path)
        return self.excel_file

    def write_in_cell_of_sheet(self, sheet, row, column, entry):
        """
        Write the <entry> value in the row, column cell of the given sheet of the excel_file object

        Parameters
        __________
        sheet : string
                the name of the sheet in which we want to write a value

        row : int
              the line number in which the value will be written

        column : int
                 the column in which the value will be written

        entry
            the value which will be written
        """
        sheet_target = self.excel_file[sheet]
        sheet_target.cell(row=row, column=column).value = entry
        self.excel_file.save(self.file_path)

    def write_list_in_line_of_sheet(self, sheet, line_number, list_entry, start_column=1):
        """
        Write the <list_entry> in the row, column cell of the given sheet of the excel_file object

        PARAMETERS
        __________
        sheet : string
                the name of the sheet in which we want to write a value

        line_number : int
                      the line number in which the value will be written (start to 1)

        list_entry : list
                     the list of values which will be written

        start_column : int
                       the column in which the line will begin (start to 1)
        """
        index = 0
        for entry in list_entry:
            self.write_in_cell_of_sheet(sheet, line_number, index + start_column, entry)
            index += 1
        self.excel_file.save(self.file_path)
